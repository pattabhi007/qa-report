import os
import sys
import argparse
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
import requests

load_dotenv()
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_EXCEL = "qa_automation_report.xlsx"
OUTPUT_CHART_DIR = "."

AUTOMATION_RUN_STATUS = {1: "Running", 2: "Success", 3: "Failure"}
VALID_PAGE_SIZES = (15, 25, 50, 100)


class TestmoClient:
    """Thin wrapper around the Testmo REST API."""

    def __init__(self, base_url, token):
        self.base_url = base_url.rstrip("/")
        if "/api/v1" not in self.base_url:
            self.base_url += "/api/v1"
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        })

    def _fetch_page(self, endpoint, params):
        resp = self.session.get(f"{self.base_url}{endpoint}", params=params)
        resp.raise_for_status()
        data = resp.json()
        items = data.get("result", []) if isinstance(data, dict) else data
        last_page = data.get("last_page", 1) if isinstance(data, dict) else 1
        return items, last_page

    def _get_all(self, endpoint, params=None):
        """Fetch every page from a paginated endpoint."""
        all_items = []
        params = params or {}
        params.setdefault("per_page", 100)
        page = 1

        while True:
            params["page"] = page
            items, last_page = self._fetch_page(endpoint, params)
            if not items:
                break
            all_items.extend(items)
            if page >= last_page:
                break
            page += 1

        return all_items

    def get_automation_sources(self, project_id):
        return self._get_all(f"/projects/{project_id}/automation/sources")

    def get_automation_runs(self, project_id, created_after=None, created_before=None):
        params = {"per_page": 100}
        if created_after:
            params["created_after"] = created_after
        if created_before:
            params["created_before"] = created_before
        return self._get_all(f"/projects/{project_id}/automation/runs", params)

    def get_automation_run(self, run_id):
        resp = self.session.get(f"{self.base_url}/automation/runs/{run_id}")
        resp.raise_for_status()
        data = resp.json()
        return data.get("result", data)


def build_source_map(client, project_id):
    """Return {source_id: source_name} mapping."""
    sources = client.get_automation_sources(project_id)
    return {s["id"]: s["name"] for s in sources}


def extract_env(tags):
    """Pull the environment name from tags like 'env-genmax-qa'."""
    for tag in (tags or []):
        if tag.startswith("env-"):
            return tag.replace("env-", "")
    return "unknown"


def extract_capability(tags):
    """Pull the capability/source name from tags like 'cap-lead-to-opportunity'."""
    for tag in (tags or []):
        if tag.startswith("cap-"):
            return tag.replace("cap-", "")
    return ""


def parse_run(run, source_map):
    """Convert a raw API run into a flat summary dict."""
    run_id = run["id"]
    tags = run.get("tags", [])
    source_id = run.get("source_id")
    source_name = source_map.get(source_id, extract_capability(tags) or str(source_id))

    passed = run.get("status1_count", 0)
    failed = run.get("failure_count", 0)
    skipped = run.get("status5_count", 0) + run.get("status6_count", 0)
    total = run.get("total_count", 0)
    pass_rate = round((passed / total) * 100, 2) if total else 0.0

    created = run.get("created_at", "")
    date_str = created[:10] if created else ""

    return {
        "RunID": run_id,
        "Run": run.get("name", ""),
        "Environment": extract_env(tags),
        "Source": source_name,
        "Date": date_str,
        "Status": AUTOMATION_RUN_STATUS.get(run.get("status", 0), "Unknown"),
        "Total": total,
        "Passed": passed,
        "Failed": failed,
        "Skipped": skipped,
        "PassRate": pass_rate,
    }


def filter_ui_runs(runs):
    """Keep only 'UI Test Run' automation runs."""
    return [r for r in runs if "UI Test Run" in r.get("name", "")]


def build_daily_trend(df, env_name):
    """Aggregate per date+source for a given environment."""
    env_df = df[df["Environment"] == env_name].copy()
    if env_df.empty:
        return pd.DataFrame()

    grouped = env_df.groupby(["Date", "Source"]).agg(
        Total=("Total", "sum"),
        Passed=("Passed", "sum"),
        Failed=("Failed", "sum"),
        Skipped=("Skipped", "sum"),
        Runs=("RunID", "count"),
    ).reset_index()
    grouped["PassRate"] = (
        (grouped["Passed"] / grouped["Total"] * 100).round(2).fillna(0)
    )
    return grouped.sort_values(["Date", "Source"])


def plot_trend(trend_df, env_name, filename):
    """Create a daily pass-rate trend chart per source."""
    if trend_df.empty:
        return

    sources = sorted(trend_df["Source"].unique())
    _, ax = plt.subplots(figsize=(14, 6))

    for source in sources:
        src_df = trend_df[trend_df["Source"] == source].sort_values("Date")
        ax.plot(src_df["Date"], src_df["PassRate"], marker="o", label=source, linewidth=2)

    ax.set_ylim(0, 105)
    ax.set_ylabel("Pass Rate (%)")
    ax.set_xlabel("Date")
    ax.set_title(f"Daily Pass Rate Trend — {env_name}")
    ax.legend(loc="lower left", fontsize=7, ncol=2)
    ax.grid(axis="y", alpha=0.3)
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_CHART_DIR, filename), dpi=150)
    plt.close()
    print(f"  Chart saved: {filename}")


def plot_stacked_bar(df, env_name, filename):
    """Create a stacked bar chart of pass/fail/skip per source for an environment."""
    env_df = df[df["Environment"] == env_name].copy()
    if env_df.empty:
        return

    by_source = env_df.groupby("Source").agg(
        Passed=("Passed", "sum"),
        Failed=("Failed", "sum"),
        Skipped=("Skipped", "sum"),
    ).sort_index()

    _, ax = plt.subplots(figsize=(12, 5))
    ax.bar(by_source.index, by_source["Passed"], label="Passed", color="#4CAF50")
    ax.bar(by_source.index, by_source["Failed"],
           bottom=by_source["Passed"], label="Failed", color="#F44336")
    ax.bar(by_source.index, by_source["Skipped"],
           bottom=by_source["Passed"] + by_source["Failed"],
           label="Skipped", color="#FF9800")
    ax.set_ylabel("Test Count")
    ax.set_title(f"Test Results by Source — {env_name}")
    ax.legend()
    plt.xticks(rotation=45, ha="right", fontsize=8)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_CHART_DIR, filename), dpi=150)
    plt.close()
    print(f"  Chart saved: {filename}")


def _latest_per_env(all_df, env_name, env_label):
    """Get the most recent run per source for a given environment."""
    env_df = all_df[all_df["Environment"] == env_name].copy()
    if env_df.empty:
        return pd.DataFrame()

    env_df = env_df.sort_values("RunID", ascending=False)
    latest = env_df.groupby("Source").first().reset_index()
    latest["Environment"] = env_label
    return latest[
        ["Source", "Environment", "RunID", "Date", "Status",
         "Total", "Passed", "Failed", "Skipped", "PassRate"]
    ]


def _build_latest_report(all_df):
    """Build a unified table: QA + Staging interleaved per source."""
    parts = []
    for env_name, label in ENV_CONFIGS:
        part = _latest_per_env(all_df, env_name, label)
        if not part.empty:
            parts.append(part)

    if not parts:
        return pd.DataFrame()

    combined = pd.concat(parts, ignore_index=True)
    combined = combined.rename(columns={"PassRate": "Pass Rate %"})
    combined = combined.sort_values(["Source", "Environment"]).reset_index(drop=True)
    return combined


ENV_CONFIGS = [
    ("genmax-qa", "QA"),
    ("genmax-staging", "Staging"),
]


NAVY = "1F3864"
ORANGE = "E36C09"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
BORDER_COLOR = "8DB4E2"
GREEN_BG = "C6EFCE"
GREEN_FG = "006100"
RED_BG = "FFC7CE"
RED_FG = "9C0006"
YELLOW_BG = "FFEB9C"
YELLOW_FG = "9C6500"

CELL_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

STYLES = {
    "navy_font": Font(name="Calibri", bold=True, color=WHITE, size=11),
    "navy_fill": PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid"),
    "orange_font": Font(name="Calibri", bold=True, color=WHITE, size=11),
    "orange_fill": PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid"),
    "cell_font": Font(name="Calibri", size=11),
    "cell_font_bold": Font(name="Calibri", size=11, bold=True),
    "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left": Alignment(horizontal="left", vertical="center"),
    "stripe": PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"),
    "green": (PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"),
              Font(name="Calibri", size=11, color=GREEN_FG, bold=True)),
    "red": (PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"),
            Font(name="Calibri", size=11, color=RED_FG, bold=True)),
    "yellow": (PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid"),
               Font(name="Calibri", size=11, color=YELLOW_FG, bold=True)),
    "passed_font": Font(name="Calibri", size=11, color=GREEN_FG, bold=True),
    "failed_font": Font(name="Calibri", size=11, color=RED_FG, bold=True),
}


def _apply_header(cell, is_orange=False):
    """Style a single header cell navy or orange."""
    key = "orange" if is_orange else "navy"
    cell.font = STYLES[f"{key}_font"]
    cell.fill = STYLES[f"{key}_fill"]
    cell.alignment = STYLES["center"]
    cell.border = CELL_BORDER


def _is_orange_header(col_name):
    """Determine if a column header is an environment or source label (orange accent)."""
    return col_name in ("source", "environment")


def _style_header_row(ws):
    """Format the header row: navy by default, orange for QA/Staging group columns."""
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        col_name = str(cell.value or "").lower()
        _apply_header(cell, is_orange=_is_orange_header(col_name))


def _style_data_cell(cell, col_name, is_stripe):
    """Format a single data cell."""
    cell.font = STYLES["cell_font"]
    cell.border = CELL_BORDER
    cell.alignment = STYLES["left"] if col_name == "source" else STYLES["center"]

    is_highlight = "pass" in col_name or "fail" in col_name or "status" in col_name
    if is_stripe and not is_highlight:
        cell.fill = STYLES["stripe"]


def _color_passrate(cell, val):
    if val >= 80:
        cell.fill, cell.font = STYLES["green"]
    elif val >= 50:
        cell.fill, cell.font = STYLES["yellow"]
    else:
        cell.fill, cell.font = STYLES["red"]


def _color_status(cell, val):
    sval = str(val).lower()
    if sval == "success":
        cell.fill, cell.font = STYLES["green"]
    elif sval == "failure":
        cell.fill, cell.font = STYLES["red"]


def _color_passed(cell, val):
    if isinstance(val, (int, float)):
        cell.font = STYLES["passed_font"]


def _color_failed(cell, val):
    if isinstance(val, (int, float)) and val > 0:
        cell.font = STYLES["failed_font"]


def _color_passrate_if_numeric(cell, val):
    if isinstance(val, (int, float)):
        _color_passrate(cell, val)


COLUMN_COLORIZERS = {
    "passrate": _color_passrate_if_numeric,
    "status": _color_status,
    "passed": _color_passed,
    "failed": _color_failed,
}


def _apply_status_color(cell, col_name):
    """Apply conditional color based on column name and cell value."""
    for keyword, colorizer in COLUMN_COLORIZERS.items():
        if keyword in col_name:
            colorizer(cell, cell.value)
            return


def _auto_column_widths(ws, df):
    """Set column widths based on content length."""
    for col_idx, col_name in enumerate(df.columns, 1):
        header_len = len(str(col_name))
        max_data = df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0
        width = min(max(header_len, max_data, 8) + 3, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _style_sheet(ws, df):
    """Apply full formatting to a worksheet."""
    _style_header_row(ws)
    col_names = [str(c.value or "").lower() for c in ws[1]]

    for row_idx in range(2, ws.max_row + 1):
        is_stripe = row_idx % 2 == 0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = col_names[col_idx - 1] if col_idx - 1 < len(col_names) else ""
            _style_data_cell(cell, col_name, is_stripe)
            _apply_status_color(cell, col_name)

    _auto_column_widths(ws, df)
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = NAVY


def _write_sheet(writer, df, sheet_name):
    """Write a styled DataFrame to a sheet."""
    if df.empty:
        return
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    _style_sheet(writer.sheets[sheet_name], df)


def _style_source_grouping(ws):
    """Merge-like visual: bold source name on first env row, lighter on second."""
    prev_source = None
    for row_idx in range(2, ws.max_row + 1):
        source_cell = ws.cell(row=row_idx, column=1)
        current_source = source_cell.value
        if current_source == prev_source:
            source_cell.font = Font(name="Calibri", size=11, color="808080")
        else:
            source_cell.font = Font(name="Calibri", size=11, bold=True)
        prev_source = current_source


def write_excel(all_df, qa_trend, staging_trend):
    """Write Excel with a single unified latest-runs sheet."""
    report = _build_latest_report(all_df)
    if report.empty:
        print("  No data to write.")
        return

    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        _write_sheet(writer, report, "Latest Runs")
        _style_source_grouping(writer.sheets["Latest Runs"])

    print(f"  Excel saved: {OUTPUT_EXCEL}")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Fetch Testmo UI Test Run results and generate trend reports."
    )
    parser.add_argument(
        "--url", default=os.environ.get("TESTMO_URL", ""),
        help="Testmo instance URL (or set TESTMO_URL env var)",
    )
    parser.add_argument(
        "--token", default=os.environ.get("TESTMO_TOKEN", ""),
        help="Testmo API token (or set TESTMO_TOKEN env var)",
    )
    parser.add_argument(
        "--project-id", default=os.environ.get("TESTMO_PROJECT_ID", "11"),
        help="Testmo project ID (default: 11)",
    )
    parser.add_argument(
        "--days", type=int, default=7,
        help="Number of days to look back (default: 7)",
    )
    parser.add_argument(
        "--run-ids", nargs="*", type=int,
        help="Specific automation run IDs (overrides --days)",
    )
    return parser.parse_args()


def fetch_runs(client, args):
    """Fetch raw automation runs from the API based on CLI arguments."""
    if args.run_ids:
        print(f"\nFetching {len(args.run_ids)} specific run(s)...")
        raw_runs = []
        for rid in args.run_ids:
            try:
                raw_runs.append(client.get_automation_run(rid))
            except requests.HTTPError as exc:
                print(f"  Warning: run {rid}: {exc}")
        return raw_runs

    since = datetime.now(timezone.utc) - timedelta(days=args.days)
    since_iso = since.strftime("%Y-%m-%dT%H:%M:%SZ")
    print(f"\nFetching automation runs since {since_iso}...")
    all_runs = client.get_automation_runs(args.project_id, created_after=since_iso)
    print(f"  Total automation runs: {len(all_runs)}")
    ui_runs = filter_ui_runs(all_runs)
    print(f"  UI Test Runs: {len(ui_runs)}")
    return ui_runs


def print_env_summary(df):
    """Print per-environment, per-source summary tables to stdout."""
    for env in sorted(df["Environment"].unique()):
        env_df = df[df["Environment"] == env]
        print(f"\n{'='*80}")
        print(f" {env.upper()}  ({len(env_df)} runs)")
        print(f"{'='*80}")
        summary = env_df.groupby("Source").agg(
            Runs=("RunID", "count"),
            Total=("Total", "sum"),
            Passed=("Passed", "sum"),
            Failed=("Failed", "sum"),
            Skipped=("Skipped", "sum"),
        ).reset_index()
        summary["PassRate"] = (summary["Passed"] / summary["Total"] * 100).round(2).fillna(0)
        print(summary.to_string(index=False))


def generate_all_outputs(df):
    """Write Excel and all charts."""
    print(f"\n{'='*80}")
    print(" Generating reports...")
    print(f"{'='*80}")

    qa_trend = build_daily_trend(df, "genmax-qa")
    staging_trend = build_daily_trend(df, "genmax-staging")

    write_excel(df, qa_trend, staging_trend)
    plot_trend(qa_trend, "GenMax QA", "trend_genmax_qa.png")
    plot_trend(staging_trend, "GenMax Staging", "trend_genmax_staging.png")
    plot_stacked_bar(df, "genmax-qa", "summary_genmax_qa.png")
    plot_stacked_bar(df, "genmax-staging", "summary_genmax_staging.png")


def main():
    args = parse_args()

    if not args.url or not args.token:
        print("Error: TESTMO_URL and TESTMO_TOKEN are required.")
        print("Set them as environment variables or pass --url and --token.")
        sys.exit(1)

    client = TestmoClient(args.url, args.token)

    print("Loading automation sources...")
    source_map = build_source_map(client, args.project_id)
    print(f"  Found {len(source_map)} sources")

    raw_runs = fetch_runs(client, args)
    if not raw_runs:
        print("No UI Test Runs found.")
        sys.exit(0)

    rows = [parse_run(r, source_map) for r in raw_runs]
    df = pd.DataFrame(rows).sort_values(["Date", "Environment", "Source"])

    print_env_summary(df)
    generate_all_outputs(df)
    print("\nDone!")


if __name__ == "__main__":
    main()
